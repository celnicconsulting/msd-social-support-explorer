# ====================IMPORTS====================
"""MSD Social Support Explorer — a visual specification in code.

This file is written to be read as the specification of the application it
draws. Every function states what appears on screen and why it appears that way,
so the page can be reconstructed from the source without running it.

How to read it
--------------
  DATA            one function per visual, returning exactly the grain that
                  visual needs. Each names the chart or table it feeds.
  SIDEBAR         the global controls, and what each one filters.
  TABS            one function per tab, whose docstring is that tab's layout,
                  block by block, in the order it renders.
  VISUALISATION   one function per chart or table: what is drawn, on which
                  axes, in which colours, and the reason behind the choice
                  wherever the choice is not obvious.
  STATIC_METHODS  shared helpers with no page of their own.
  MAIN            page assembly.

The page
--------
  hazard-striped provenance banner                        render_header
  sidebar: period, geography, benefit, ethnicity basis     render_sidebar
  nine tabs                                               render_main_tabs
    Overview            how many people are on a main benefit, and who
    Map                 where they are
    Housing             emergency housing and the social housing register
    Hardship            supplementary and hardship assistance
    StudyLink           student support against total tertiary participation
    Retirement income   NZ Super and the Veteran's Pension, and their cost
    All assistance      every programme's spend on one axis
    Pipeline            where every figure came from
    Build Notes         how the platform was built, from its own write-up
  sources and licence expander                            render_attribution

Rules every visual on this page obeys
-------------------------------------
  - A control never offers a period, area or benefit its charts cannot answer:
    each picker is built from the series it filters.
  - Counts, dollars and percentages never share an axis. VALUE_KIND travels
    with every row, so a chart filters on the data rather than on the wording
    of a label.
  - Suppressed cells stay blank and are captioned as suppressions, never zero.
  - Total-response ethnicity is never summed, and the chart says which basis it
    is drawing.
  - Treasury outturn and forecast use different line styles and are separated
    by a marked boundary; they are never one series.
  - Overlapping populations are drawn as lines, never stacked or totalled.
  - Every table downloads through build_styled_excel, so exports match wherever
    they came from.

Built to the WCC Snowflake Streamlit conventions: the same section separators,
@st.cache_data query methods, render_* visual methods and a thin main(). The data
layer reads a local DuckDB mart instead of a Snowpark session, so the structure
ports to Snowflake Streamlit by swapping run_query() for session.sql().
"""
import io
import os
import re

import duckdb
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pydeck as pdk
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", page_title="Social Support Statistics", page_icon="🧭")

# ====================SESSION====================
# Local DuckDB stands in for get_active_session(). Everything downstream goes
# through run_query(), which is the only line that changes when this app is
# deployed to Snowflake Streamlit.
DB_PATH = os.environ.get(
    "MSD_DB_PATH",
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                 "data", "msd_platform_public.duckdb"),
)
df_db_schema = "MSD_MART"

# Markdown carried as files rather than strings. The build write-up is
# documentation of this application, edited on its own; embedding a copy in this
# module would give it a second version free to drift from the first. Folders are
# searched in order and the first hit wins: an explicit override, the public
# repository's reference folder, then the platform root the working copy sits in.
REFERENCE_DIRS = [
    os.environ.get("MSD_REFERENCE_DIR", ""),
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reference"),
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
]

# A write-up is named for the module it documents: <module>__readme.md. Derived
# rather than typed, so renaming this file renames the document it looks for.
BUILD_NOTES_DOC = os.path.splitext(os.path.basename(__file__))[0] + "__readme.md"

# PALETTE is the ordered sequence for categories with no fixed identity, so any
# chart drawing an arbitrary set of series looks like every other one.
PALETTE = ["#2E86AB", "#E4572E", "#17BEBB", "#F6AE2D", "#7B6D8D", "#4C9F70", "#C33C54"]

# BENEFIT_COLOURS pins the named benefit groups instead, so a group keeps the
# same colour on the Overview area chart, the monthly lines and the snapshot,
# and a reader can track one benefit across the whole page.
BENEFIT_COLOURS = {
    "Jobseeker Support": "#2E86AB",
    "Sole Parent Support": "#E4572E",
    "Supported Living Payment": "#4C9F70",
    "Other main benefits": "#F6AE2D",
    "Youth Payment / Young Parent Payment": "#7B6D8D",
    "All main benefits": "#1A3A5C",
}


@st.cache_resource
def get_connection():
    """Read-only DuckDB handle, cached for the session.

    Stands in for get_active_session(). Nothing else in the file opens a
    connection.
    """
    return duckdb.connect(DB_PATH, read_only=True)


@st.cache_data(show_spinner=False)
def run_query(sql, params=None):
    """Single point of database access. Swap for session.sql(...).to_pandas() on Snowflake."""
    con = get_connection()
    return con.execute(sql, params or []).df()


# ====================DATA====================
# One query per visual, at exactly the grain that visual needs. Nothing here
# reshapes data for presentation: the mart already carries VALUE_KIND,
# ETHNICITY_BASIS, IS_SUPPRESSED, IS_PREFERRED_SOURCE and the H3 cells, so a
# chart filters on columns rather than on the wording of a label. Every
# docstring names the chart or table the result feeds.
@st.cache_data(show_spinner=False)
def get_periods(df_db_schema, period_type):
    """Every period of one type, from the conformed date dimension.

    Used only where a control needs the full published calendar — the emergency
    housing quarter picker. Every other picker is built from its own series.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, PERIOD_DATE, CALENDAR_YEAR
        FROM {s}.DIM_PERIOD
        WHERE PERIOD_TYPE = ?
        ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [period_type])


@st.cache_data(show_spinner=False)
def get_quarter_options(df_db_schema):
    """Quarters the benefit series actually covers.

    DIM_PERIOD spans every source, including Treasury's monthly forecast back to
    1996, so driving a control from it offers periods the charts cannot answer.
    Each picker is built from the series it filters.
    """
    return run_query(
        """
        SELECT DISTINCT PERIOD, PERIOD_SORT FROM {s}.VW_BENEFIT_TOTALS
        ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_month_options(df_db_schema):
    """Months the monthly benefit fact covers — the sidebar "Latest month" picker.

    Built from the fact, not DIM_PERIOD, for the same reason as the quarter
    slider above.
    """
    return run_query(
        """
        SELECT DISTINCT PERIOD, PERIOD_SORT FROM {s}.FACT_BENEFIT_MONTHLY
        WHERE IS_PREFERRED_SOURCE ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_benefit_groups(df_db_schema):
    """Working-age benefit groups for the sidebar multiselect.

    The "All main benefits" rollup is excluded because it would double count
    against its own components on a stacked chart, and the two pensions because
    they are not working-age support and have their own tab.
    """
    return run_query(
        """
        SELECT DISTINCT BENEFIT_GROUP_STD AS BENEFIT_GROUP
        FROM {s}.VW_BENEFIT_TOTALS
        WHERE BENEFIT_GROUP_STD NOT IN ('All main benefits', 'New Zealand Superannuation',
                                        'Veteran''s Pension')
        ORDER BY 1
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_geographies(df_db_schema, geo_level):
    """Area names at one level, with the indicative centroid the map plots."""
    return run_query(
        """
        SELECT GEO_NAME, LATITUDE, LONGITUDE, HAS_COORDINATES
        FROM {s}.DIM_GEOGRAPHY WHERE GEO_LEVEL = ? ORDER BY GEO_NAME
        """.format(s=df_db_schema), [geo_level])


@st.cache_data(show_spinner=False)
def get_national_trend(df_db_schema, p_from, p_to, benefit_groups):
    """Quarterly working-age recipients by benefit group, national.

    Reads the block Total row via VW_BENEFIT_TOTALS rather than summing a
    characteristic: ethnicity is total response from December 2021 and summing it
    would double count people who identify with more than one group.
    """
    ph = ", ".join(["?"] * len(benefit_groups))
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, BENEFIT_GROUP_STD AS BENEFIT_GROUP,
               SUM(CLIENT_COUNT) AS CLIENT_COUNT
        FROM {s}.VW_BENEFIT_TOTALS
        WHERE GEO_LEVEL = 'NATIONAL' AND BENEFIT_GROUP_STD IN ({ph})
          AND PERIOD_SORT BETWEEN ? AND ?
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema, ph=ph), list(benefit_groups) + [p_from, p_to])


@st.cache_data(show_spinner=False)
def get_monthly_trend(df_db_schema, p_from, p_to):
    """Monthly recipients by headline benefit — the Overview line chart on the right."""
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, LABEL_1 AS BENEFIT, SUM(VALUE) AS CLIENT_COUNT
        FROM {s}.FACT_BENEFIT_MONTHLY
        WHERE IS_PREFERRED_SOURCE AND VALUE_KIND = 'COUNT'
          AND PERIOD_SORT BETWEEN ? AND ?
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [p_from, p_to])


@st.cache_data(show_spinner=False)
def get_characteristic_breakdown(df_db_schema, period, geo_level, geo_names,
                                 benefit_groups, characteristic):
    """Recipients split by one characteristic — the three Overview bar charts.

    ETHNICITY_BASIS travels with the rows so the chart can caption which basis
    it is drawing, and warn when the bars must not be summed.
    """
    gph = ", ".join(["?"] * len(geo_names)) if geo_names else "''"
    bph = ", ".join(["?"] * len(benefit_groups))
    return run_query(
        """
        SELECT GEO_NAME, BENEFIT_GROUP, LABEL_2 AS CHARACTERISTIC_VALUE,
               SUM(VALUE) AS CLIENT_COUNT, ANY_VALUE(ETHNICITY_BASIS) AS ETHNICITY_BASIS
        FROM {s}.FACT_BENEFIT_CHARACTERISTIC
        WHERE IS_PREFERRED_SOURCE AND VALUE_KIND = 'COUNT'
          AND PERIOD = ? AND GEO_LEVEL = ? AND LABEL_1 = ?
          AND GEO_NAME IN ({gph}) AND BENEFIT_GROUP IN ({bph})
          AND LABEL_2 IS NOT NULL
        GROUP BY ALL ORDER BY CLIENT_COUNT DESC
        """.format(s=df_db_schema, gph=gph, bph=bph),
        [period, geo_level, characteristic] + list(geo_names) + list(benefit_groups))


@st.cache_data(show_spinner=False)
def get_latest_snapshot(df_db_schema, period, geo_level, benefit_groups):
    """One quarter by area and benefit group — the Overview snapshot table and its export."""
    bph = ", ".join(["?"] * len(benefit_groups))
    return run_query(
        """
        SELECT GEO_NAME AS GEOGRAPHY, BENEFIT_GROUP_STD AS BENEFIT_GROUP,
               SUM(CLIENT_COUNT) AS CLIENT_COUNT
        FROM {s}.VW_BENEFIT_TOTALS
        WHERE PERIOD = ? AND GEO_LEVEL = ? AND BENEFIT_GROUP_STD IN ({bph})
        GROUP BY ALL ORDER BY CLIENT_COUNT DESC
        """.format(s=df_db_schema, bph=bph), [period, geo_level] + list(benefit_groups))


@st.cache_data(show_spinner=False)
def get_map_data(df_db_schema, period, geo_level, benefit_groups):
    """Client counts joined to indicative centroids and pre-computed H3 cells.

    Feeds the hexagon layer on the Map tab at region, regional council and
    Auckland local board level. The H3 cells are computed once in the mart, so
    the app does no geometry at render time.
    """
    bph = ", ".join(["?"] * len(benefit_groups))
    return run_query(
        """
        SELECT f.GEO_NAME, SUM(f.CLIENT_COUNT) AS CLIENT_COUNT,
               g.LATITUDE, g.LONGITUDE,
               g.H3_RES_3, g.H3_RES_4, g.H3_RES_5, g.H3_RES_6, g.H3_RES_7
        FROM {s}.VW_BENEFIT_TOTALS f
        JOIN {s}.DIM_GEOGRAPHY g
          ON g.GEO_LEVEL = f.GEO_LEVEL AND g.GEO_NAME = f.GEO_NAME
        WHERE f.PERIOD = ? AND f.GEO_LEVEL = ?
          AND f.BENEFIT_GROUP_STD IN ({bph}) AND g.HAS_COORDINATES
        GROUP BY ALL HAVING SUM(f.CLIENT_COUNT) > 0 ORDER BY CLIENT_COUNT DESC
        """.format(s=df_db_schema, bph=bph), [period, geo_level] + list(benefit_groups))


@st.cache_data(show_spinner=False)
def get_map_data_ta(df_db_schema, period):
    """TA-level monthly counts — the Map tab's default level.

    The quarterly fact sheets do not go below TA totals, so this is the only
    series fine enough to hexagon at territorial authority level.
    """
    return run_query(
        """
        SELECT f.GEO_NAME, f.MEASURE, SUM(f.VALUE) AS CLIENT_COUNT,
               g.LATITUDE, g.LONGITUDE,
               g.H3_RES_3, g.H3_RES_4, g.H3_RES_5, g.H3_RES_6, g.H3_RES_7
        FROM {s}.FACT_BENEFIT_MONTHLY_GEO f
        JOIN {s}.DIM_GEOGRAPHY g
          ON g.GEO_LEVEL = f.GEO_LEVEL AND g.GEO_NAME = f.GEO_NAME
        WHERE f.IS_PREFERRED_SOURCE AND f.VALUE_KIND = 'COUNT'
          AND f.PERIOD = ? AND f.GEO_LEVEL = 'TA' AND g.HAS_COORDINATES
        GROUP BY ALL HAVING SUM(f.VALUE) > 0 ORDER BY CLIENT_COUNT DESC
        """.format(s=df_db_schema), [period])


@st.cache_data(show_spinner=False)
def get_emergency_housing_trend(df_db_schema):
    """Monthly emergency housing grants and dollars — the two charts at the top of Housing.

    VALUE_KIND travels with the rows because counts and dollars share the
    worksheet; the charts filter on it rather than on the wording of the label.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, LABEL_1 AS METRIC, VALUE_KIND, SUM(VALUE) AS VALUE
        FROM {s}.FACT_HOUSING_EMERGENCY
        WHERE IS_PREFERRED_SOURCE AND LABEL_1 IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_register_trend(df_db_schema):
    """Housing Register and Transfer Register applications — the Housing line chart.

    Summed from the territorial authority block, with the totals row and the
    suppression notes excluded so nothing is counted twice.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, MEASURE AS REGISTER, SUM(VALUE) AS VALUE
        FROM {s}.FACT_HOUSING_REGISTER_QUARTERLY
        WHERE IS_PREFERRED_SOURCE AND VALUE_KIND = 'COUNT'
          AND MEASURE IS NOT NULL
          AND SECTION ILIKE '%Territorial Authority%'
          AND LABEL_1 NOT ILIKE '%suppressed%' AND LABEL_1 NOT ILIKE '%total%'
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_eh_by_ta(df_db_schema, period):
    """Emergency housing by territorial authority, one quarter — the Housing detail table.

    IS_SUPPRESSED is carried through so the table can state that a blank cell is
    a confidentiality suppression rather than a zero.
    """
    return run_query(
        """
        SELECT GEO_NAME AS TERRITORIAL_AUTHORITY, SECTION AS MEASURE_GROUP,
               LABEL_1 AS MEASURE, VALUE, VALUE_KIND, IS_SUPPRESSED
        FROM {s}.FACT_EMERGENCY_HOUSING_TA
        WHERE IS_PREFERRED_SOURCE AND PERIOD = ?
        ORDER BY TERRITORIAL_AUTHORITY, MEASURE_GROUP, MEASURE
        """.format(s=df_db_schema), [period])


@st.cache_data(show_spinner=False)
def get_eh_duration_matrix(df_db_schema):
    """Households by weeks-in-emergency-housing band — the Housing heatmap.

    The average-duration row is excluded: it shares the band labels but is a
    different measure entirely.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, LABEL_1 AS DURATION_BAND, SUM(VALUE) AS HOUSEHOLDS
        FROM {s}.FACT_HOUSING_EMERGENCY
        WHERE IS_PREFERRED_SOURCE AND VALUE_KIND = 'COUNT'
          AND regexp_matches(LOWER(LABEL_1), 'weeks in eh')
          AND NOT regexp_matches(LOWER(LABEL_1), 'average')
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_hardship_trend(df_db_schema):
    """Monthly supplementary and hardship assistance — the two Hardship charts.

    SECTION separates point-in-time recipients from counts of everyone helped
    during the month, which the two charts keep apart.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, SECTION, LABEL_1 AS ASSISTANCE_TYPE,
               VALUE_KIND, SUM(VALUE) AS VALUE
        FROM {s}.FACT_HARDSHIP_MONTHLY
        WHERE IS_PREFERRED_SOURCE AND LABEL_1 IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_food_grants(df_db_schema):
    """Food Special Needs Grants by regional council — the two food grant charts."""
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, GEO_NAME AS REGIONAL_COUNCIL,
               VALUE_KIND, SUM(VALUE) AS VALUE
        FROM {s}.FACT_FOOD_GRANTS_REGION
        WHERE IS_PREFERRED_SOURCE AND GEO_NAME IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_supplementary_region(df_db_schema, p_from, p_to):
    """Supplementary assistance by Work and Income region — the Hardship area chart and table."""
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, GEO_NAME AS WI_REGION, LABEL_1 AS ASSISTANCE_TYPE,
               VALUE_KIND, SUM(VALUE) AS VALUE
        FROM {s}.FACT_SUPPLEMENTARY_REGION
        WHERE IS_PREFERRED_SOURCE AND GEO_NAME IS NOT NULL
          AND PERIOD_SORT BETWEEN ? AND ?
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [p_from, p_to])


@st.cache_data(show_spinner=False)
def get_studylink_windows(df_db_schema):
    """Year-to-date windows published — the StudyLink "Reporting window" picker."""
    return run_query(
        """
        SELECT DISTINCT WINDOW_MONTHS FROM {s}.VW_STUDYLINK_ANNUAL ORDER BY 1
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_studylink_annual(df_db_schema, window_months):
    """Recipients and dollars on one year-to-date window.

    StudyLink restates the whole timeseries each quarter, truncated to the
    release quarter, so every year in a given window is directly comparable and
    windows must never be mixed.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, PRODUCT, RECIPIENTS, AMOUNT_NZD
        FROM {s}.VW_STUDYLINK_ANNUAL
        WHERE WINDOW_MONTHS = ? ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [window_months])


@st.cache_data(show_spinner=False)
def get_studylink_provider(df_db_schema, window_months):
    """StudyLink recipients and dollars by provider type — the two provider charts."""
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, PRODUCT, PROVIDER_TYPE, PROVIDER_TYPE_STD,
               RECIPIENTS, AMOUNT_NZD
        FROM {s}.VW_STUDYLINK_PROVIDER
        WHERE WINDOW_MONTHS = ? AND RECIPIENTS IS NOT NULL
        ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [window_months])


@st.cache_data(show_spinner=False)
def _has_tertiary(df_db_schema):
    """Whether the Ministry of Education extract is present in this mart.

    The trimmed public build can ship without it, so every tertiary visual
    degrades to StudyLink-only rather than erroring.
    """
    return run_query(
        """
        SELECT COUNT(*) AS N FROM information_schema.tables
        WHERE table_schema = ? AND table_name = 'VW_TERTIARY_PARTICIPATION'
        """, [df_db_schema]).N.iloc[0] > 0


@st.cache_data(show_spinner=False)
def get_tertiary_total(df_db_schema, measure):
    """All tertiary students (or EFTS) per year, Ministry of Education.

    Returns empty when the Education Counts extract has not been loaded, so the
    StudyLink tab degrades to StudyLink-only rather than failing.
    """
    if not _has_tertiary(df_db_schema):
        return pd.DataFrame(columns=["PERIOD", "PERIOD_SORT", "VALUE"])
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, VALUE
        FROM {s}.VW_TERTIARY_PARTICIPATION
        WHERE MEASURE = ? AND QUALIFICATION_LEVEL = 'Total'
          AND STUDENT_ORIGIN = 'Total' AND SUB_SECTOR = 'Total'
        ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [measure])


@st.cache_data(show_spinner=False)
def get_tertiary_by_provider(df_db_schema, measure):
    """Tertiary students by provider type — the denominator of the support share chart.

    MSD_PROVIDER_TYPE is the mart's crosswalk from the Ministry's subsectors to
    MSD's provider categories, which is what makes the two divisible at all.
    """
    if not _has_tertiary(df_db_schema):
        return pd.DataFrame(columns=["PERIOD", "PERIOD_SORT", "PROVIDER_TYPE_STD",
                                     "MSD_PROVIDER_TYPE", "VALUE"])
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, PROVIDER_TYPE_STD, MSD_PROVIDER_TYPE, SUM(VALUE) AS VALUE
        FROM {s}.VW_TERTIARY_PARTICIPATION
        WHERE MEASURE = ? AND QUALIFICATION_LEVEL = 'Total'
          AND STUDENT_ORIGIN = 'Total' AND MSD_PROVIDER_TYPE IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema), [measure])


@st.cache_data(show_spinner=False)
def get_studylink(df_db_schema):
    """Raw StudyLink fact at full grain.

    Not wired to a visual: the tab reads the semantic views above, which already
    resolve the reporting window. Kept as the escape hatch for questions the
    views do not answer.
    """
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, MEASURE AS PRODUCT, SECTION, LABEL_1, LABEL_2,
               VALUE_KIND, SUM(VALUE) AS VALUE
        FROM {s}.FACT_STUDYLINK
        WHERE IS_PREFERRED_SOURCE AND MEASURE IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_pipeline_gaps(df_db_schema):
    """Expected against actual releases per series — the coverage heatmap and two of its KPIs."""
    return run_query(
        """
        SELECT SCHEMA_NAME, FAMILY, PERIOD_TYPE, CADENCE_MONTHS, PERIOD, IS_PRESENT,
               COVERAGE_START, COVERAGE_END, RELEASE_COUNT
        FROM {s}.META_DATA_GAPS ORDER BY SCHEMA_NAME, FAMILY, PERIOD
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_pipeline_manifest(df_db_schema):
    """Every downloaded source file with size, date and URL — the Pipeline manifest table."""
    return run_query(
        """
        SELECT DATASET_ID, PERIOD, PERIOD_TYPE, FILE_NAME, FILE_EXT, FILE_SIZE,
               DOWNLOADED_AT, SOURCE_URL
        FROM {s}.META_DOWNLOAD_MANIFEST ORDER BY DATASET_ID, PERIOD, FILE_NAME
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_pipeline_coverage(df_db_schema):
    """Staged records per fact table — the Pipeline staging coverage table."""
    return run_query(
        """
        SELECT FACT_NAME, COUNT(*) AS SOURCE_TABLES, SUM(RECORDS) AS STAGED_RECORDS,
               MIN(MIN_PERIOD) AS FIRST_PERIOD, MAX(MAX_PERIOD) AS LAST_PERIOD
        FROM {s}.META_STAGING_INVENTORY GROUP BY ALL ORDER BY STAGED_RECORDS DESC
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_fact_catalog(df_db_schema):
    """Row counts and period span per mart fact — the "Mart fact rows" KPI and the provenance note."""
    return run_query("SELECT * FROM {s}.META_FACT_CATALOG ORDER BY ROWS DESC".format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_skipped_sources(df_db_schema):
    """Raw worksheets deliberately not staged, with the reason — the Pipeline expander."""
    return run_query(
        """
        SELECT SCHEMA_NAME, TABLE_NAME, REASON
        FROM {s}.META_STAGING_SKIPPED ORDER BY SCHEMA_NAME, TABLE_NAME
        """.format(s=df_db_schema))



@st.cache_data(show_spinner=False)
def get_retirement_summary(df_db_schema):
    """NZ Superannuation and Veteran's Pension recipients, quarterly."""
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, PENSION, SUM(RECIPIENTS) AS RECIPIENTS
        FROM {s}.VW_RETIREMENT_INCOME
        WHERE SOURCE_TABLE ILIKE '%BENEFIT_TYPE%'
          AND PENSION <> 'New Zealand Superannuation and Veteran''s Pension'
        GROUP BY ALL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_retirement_characteristics(df_db_schema, period, characteristic):
    """One characteristic split for both pensions — the Retirement stacked bars.

    Split by pension, because New Zealand Superannuation and the Veteran's
    Pension share row labels on the same worksheet and would otherwise merge.
    """
    return run_query(
        """
        SELECT PENSION, CHARACTERISTIC_VALUE, SUM(RECIPIENTS) AS RECIPIENTS
        FROM {s}.VW_RETIREMENT_INCOME
        WHERE PERIOD = ? AND CHARACTERISTIC = ? AND CHARACTERISTIC_VALUE IS NOT NULL
          AND SOURCE_TABLE ILIKE '%RECIPIENT_CHARS%'
        GROUP BY ALL ORDER BY RECIPIENTS DESC
        """.format(s=df_db_schema), [period, characteristic])


@st.cache_data(show_spinner=False)
def get_retirement_characteristic_types(df_db_schema, period):
    """Characteristics published for one quarter — decides which bar charts appear."""
    return run_query(
        """
        SELECT DISTINCT CHARACTERISTIC FROM {s}.VW_RETIREMENT_INCOME
        WHERE PERIOD = ? AND CHARACTERISTIC IS NOT NULL AND CHARACTERISTIC_VALUE IS NOT NULL
          AND SOURCE_TABLE ILIKE '%RECIPIENT_CHARS%'
        ORDER BY 1
        """.format(s=df_db_schema), [period])


@st.cache_data(show_spinner=False)
def get_retirement_ta(df_db_schema, period, pension):
    """Pension recipients by territorial authority, with H3 cells — the Retirement map."""
    return run_query(
        """
        SELECT GEO_NAME, SUM(RECIPIENTS) AS RECIPIENTS,
               ANY_VALUE(LATITUDE) AS LATITUDE, ANY_VALUE(LONGITUDE) AS LONGITUDE,
               ANY_VALUE(H3_RES_3) AS H3_RES_3, ANY_VALUE(H3_RES_4) AS H3_RES_4,
               ANY_VALUE(H3_RES_5) AS H3_RES_5, ANY_VALUE(H3_RES_6) AS H3_RES_6,
               ANY_VALUE(H3_RES_7) AS H3_RES_7
        FROM {s}.VW_RETIREMENT_INCOME_TA
        WHERE PERIOD = ? AND PENSION = ?
        GROUP BY ALL HAVING SUM(RECIPIENTS) > 0 ORDER BY RECIPIENTS DESC
        """.format(s=df_db_schema), [period, pension])


@st.cache_data(show_spinner=False)
def get_retirement_periods(df_db_schema):
    """Quarters with a territorial authority breakdown — the Retirement map quarter picker."""
    return run_query(
        """
        SELECT DISTINCT PERIOD, PERIOD_SORT FROM {s}.VW_RETIREMENT_INCOME_TA ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_expense_line(df_db_schema, lines):
    """Treasury expenditure for named benefit lines, actual and forecast.

    BASIS comes back with the rows so a chart can draw outturn and projection in
    different line styles instead of joining them into one misleading series.
    MSD publishes no expenditure at all, which is why cost comes from Treasury.
    """
    ph = ", ".join(["?"] * len(lines))
    return run_query(
        """
        SELECT PERIOD, PERIOD_SORT, EXPENSE_LINE, BASIS, AMOUNT_NZD_MILLION
        FROM {s}.VW_GOVT_EXPENSE
        WHERE TREASURY_TABLE ILIKE '%5.2%' AND EXPENSE_LINE IN ({ph})
        ORDER BY PERIOD_SORT
        """.format(s=df_db_schema, ph=ph), list(lines))


@st.cache_data(show_spinner=False)
def get_all_assistance(df_db_schema):
    """Every assistance programme's spend and recipients — the All assistance tab.

    BASIS separates Treasury outturn from forecast so the area chart can mark
    where one ends and the other begins.
    """
    return run_query(
        """
        SELECT PROGRAMME, PERIOD, PERIOD_SORT, BASIS, SPEND_SOURCE,
               AMOUNT_NZD_MILLION, RECIPIENTS
        FROM {s}.VW_ALL_ASSISTANCE
        WHERE AMOUNT_NZD_MILLION IS NOT NULL ORDER BY PERIOD_SORT
        """.format(s=df_db_schema))


@st.cache_data(show_spinner=False)
def get_people_supported(df_db_schema):
    """Recipients across every programme, on a common December-quarter year.

    Benefits and pensions are point-in-time quarterly counts; StudyLink is a
    full-year count of everyone supported at any time. They are shown together
    but never added, because a person can appear in more than one.
    """
    return run_query(
        """
        SELECT CAST(SUBSTR(PERIOD, 1, 4) AS VARCHAR) AS YEAR, 'Main benefits' AS PROGRAMME,
               SUM(CLIENT_COUNT) AS RECIPIENTS
        FROM {s}.VW_BENEFIT_TOTALS
        WHERE GEO_LEVEL = 'NATIONAL' AND BENEFIT_GROUP_STD = 'All main benefits'
          AND PERIOD LIKE '%Q4'
        GROUP BY ALL
        UNION ALL
        SELECT CAST(SUBSTR(PERIOD, 1, 4) AS VARCHAR), PENSION, SUM(RECIPIENTS)
        FROM {s}.VW_RETIREMENT_INCOME
        WHERE SOURCE_TABLE ILIKE '%BENEFIT_TYPE%' AND PERIOD LIKE '%Q4'
          AND PENSION <> 'New Zealand Superannuation and Veteran''s Pension'
        GROUP BY ALL
        UNION ALL
        SELECT PERIOD, CASE WHEN PRODUCT = 'Student Allowance' THEN 'Student Allowance'
                            ELSE 'Student Loans' END, MAX(RECIPIENTS)
        FROM {s}.VW_STUDYLINK_ANNUAL
        WHERE WINDOW_MONTHS = 12 AND RECIPIENTS IS NOT NULL
        GROUP BY ALL
        ORDER BY 1
        """.format(s=df_db_schema))

@st.cache_data(show_spinner=False)
def get_reference_doc(file_name):
    """Load a markdown reference document from disk, verbatim — the Build Notes tab.

    Returns the text, or None when no copy is on the search path, so a
    deployment shipped without the reference folder renders a short notice
    rather than failing.
    """
    for folder in REFERENCE_DIRS:
        if not folder:
            continue
        path = os.path.join(folder, file_name)
        if os.path.exists(path):
            with io.open(path, encoding="utf-8") as fh:
                return fh.read()
    return None


# ====================SIDEBAR====================
# The only global state. Filters chosen here travel to every tab as a plain
# dict, so no tab reads a widget belonging to another.
def render_sidebar():
    """Global controls, rendered once and handed to every tab as a dict.

    Top to bottom:
      title and provenance caption
      Period          quarter range slider, defaulting to the last 25 quarters
                      latest-month selectbox for the monthly series
      Geography       level selectbox, then an area multiselect for that level
      Benefit         benefit group multiselect
      Ethnicity basis caption and the December 2021 split toggle
      a standing provenance note

    Every picker is built from the series it filters, so no control can offer a
    period, area or benefit the charts cannot answer.
    """
    st.sidebar.title("🧭 Social Support Statistics")
    st.sidebar.caption(
        "Derived from publicly available data: Ministry of Social Development "
        "(including StudyLink), Ministry of Education and the Treasury."
    )

    quarters = get_quarter_options(df_db_schema)
    qlist = quarters.PERIOD.tolist()

    st.sidebar.header("Period")
    default_from = qlist[max(0, len(qlist) - 25)]
    q_from, q_to = st.sidebar.select_slider(
        "Quarter range", options=qlist, value=(default_from, qlist[-1]))

    months = get_month_options(df_db_schema)
    mlist = months.PERIOD.tolist()
    m_focus = st.sidebar.selectbox("Latest month (monthly series)", options=mlist[::-1], index=0)

    st.sidebar.header("Geography")
    geo_level = st.sidebar.selectbox(
        "Level",
        options=["NATIONAL", "WI_REGION", "REGIONAL_COUNCIL", "TA", "AUCKLAND_BOARD"],
        format_func=lambda x: {"NATIONAL": "National", "WI_REGION": "Work and Income region",
                               "REGIONAL_COUNCIL": "Regional council", "TA": "Territorial authority",
                               "AUCKLAND_BOARD": "Auckland local board"}[x],
        index=1)
    geos = get_geographies(df_db_schema, geo_level)
    geos = geos[geos.HAS_COORDINATES] if geos.HAS_COORDINATES.any() else geos
    geo_names = geos.GEO_NAME.tolist()
    selected_geos = st.sidebar.multiselect("Areas", options=geo_names, default=geo_names)

    st.sidebar.header("Benefit")
    bgs = get_benefit_groups(df_db_schema).BENEFIT_GROUP.tolist()
    selected_bgs = st.sidebar.multiselect("Benefit groups", options=bgs, default=bgs)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Ethnicity basis")
    st.sidebar.caption(
        "MSD changed basis in the December 2021 quarter: prioritised ethnicity before, "
        "total response after. The two series are not comparable and total response "
        "must never be summed to a total — a person counts in every group they name."
    )
    split_ethnicity = st.sidebar.checkbox("Split ethnicity charts at Dec 2021", value=True)

    st.sidebar.markdown("---")
    st.sidebar.info(
        "**Every figure on this page is published New Zealand government data**, from "
        "the Ministry of Social Development (benefits, housing, hardship and StudyLink), "
        "the Ministry of Education (Education Counts tertiary participation) and the "
        "Treasury (core Crown expense tables). No synthetic or modelled values are used. "
        "Suppressed cells stay blank rather than becoming zero. Map markers are "
        "indicative area centroids, not boundaries. Every source file is listed in the "
        "Pipeline tab."
    )

    qsort = dict(zip(quarters.PERIOD, quarters.PERIOD_SORT))
    msort = dict(zip(months.PERIOD, months.PERIOD_SORT))
    return {
        "q_from": q_from, "q_to": q_to,
        "q_from_sort": int(qsort[q_from]), "q_to_sort": int(qsort[q_to]),
        "m_focus": m_focus, "m_focus_sort": int(msort[m_focus]),
        "geo_level": geo_level, "geos": selected_geos, "benefit_groups": selected_bgs,
        "split_ethnicity": split_ethnicity, "all_quarters": qlist,
    }


# ====================TABS====================
def render_main_tabs(f):
    """Nine tabs, left to right, each answering one question.

      Overview           how many people are on a main benefit, and who they are
      Map                where they are
      Housing            emergency housing and the social housing register
      Hardship           supplementary and hardship assistance
      StudyLink          student support against total tertiary participation
      Retirement income  NZ Super and the Veteran's Pension, and what they cost
      All assistance     every programme's spend on one axis
      Pipeline           where every figure came from
      Build Notes        how the platform was built, from its own write-up

    `f` is the filter dict returned by render_sidebar.
    """
    t1, t2, t3, t4, t5, t6, t7, t8, t9 = st.tabs(
        ["📊 Overview", "🗺️ Map", "🏠 Housing", "💵 Hardship", "🎓 StudyLink",
         "🧓 Retirement income", "🏛️ All assistance", "⚙️ Pipeline", "📐 Build Notes"])
    with t1:
        render_tab_overview(f)
    with t2:
        render_tab_map(f)
    with t3:
        render_tab_housing(f)
    with t4:
        render_tab_hardship(f)
    with t5:
        render_tab_studylink(f)
    with t6:
        render_tab_retirement(f)
    with t7:
        render_tab_all_assistance(f)
    with t8:
        render_tab_pipeline(f)
    with t9:
        render_tab_build_notes(f)


def render_tab_overview(f):
    """Overview — how many people are on a main benefit.

      four KPI tiles: latest count, quarter on quarter, same quarter a year
        earlier, quarters shown
      ---
      3:2 columns: stacked area by benefit group | monthly headline lines
      ---
      three equal columns: age, gender and ethnic group bars for the closing quarter
      ---
      snapshot table by area, with an Excel export

    An empty selection short-circuits to a warning rather than an empty chart.
    """
    st.header("Working-age main benefit recipients")
    if not f["benefit_groups"]:
        st.warning("Select at least one benefit group in the sidebar.")
        return

    trend = get_national_trend(df_db_schema, f["q_from_sort"], f["q_to_sort"], f["benefit_groups"])
    if trend.empty:
        st.warning("No data for the selected period and benefit groups.")
        return

    render_overview_kpis(trend)
    st.markdown("---")

    c1, c2 = st.columns([3, 2])
    with c1:
        render_benefit_area_chart(trend)
    with c2:
        render_monthly_headline(f)

    st.markdown("---")
    render_characteristic_charts(f)
    st.markdown("---")
    render_snapshot_table(f)


def render_tab_map(f):
    """Map — where recipients are.

      caption stating that hexagons mark centroids, not boundaries
      three equal columns: map level | hexagon resolution | jump-to city
      extruded H3 hexagon layer, yellow to red by count
      ---
      area detail table with an Excel export, H3 columns dropped

    Territorial authority reads the monthly fact, the only series published that
    deep; every other level reads the quarterly one, so the subtitle names which.
    """
    st.header("Where recipients are")
    st.caption(
        "Hexagons sit on an indicative centroid for each area. MSD publishes counts by "
        "area name with no boundaries, so a hexagon marks a place, not its extent."
    )

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        map_level = st.selectbox(
            "Map level", ["TA", "WI_REGION", "REGIONAL_COUNCIL", "AUCKLAND_BOARD"],
            format_func=lambda x: {"TA": "Territorial authority (monthly)",
                                   "WI_REGION": "Work and Income region (quarterly)",
                                   "REGIONAL_COUNCIL": "Regional council (quarterly)",
                                   "AUCKLAND_BOARD": "Auckland local board (quarterly)"}[x],
            key="map_level")
    with c2:
        resolution = st.select_slider("Hexagon resolution", options=[3, 4, 5, 6, 7], value=4,
                                      key="map_res")
    with c3:
        jump = st.selectbox("Jump to", ["New Zealand", "Auckland", "Wellington", "Christchurch",
                                        "Hamilton", "Dunedin"], key="map_jump")

    if map_level == "TA":
        data = get_map_data_ta(df_db_schema, f["m_focus"])
        subtitle = "Jobseeker Support, other main benefits and Accommodation Supplement, %s" % f["m_focus"]
        data = data.groupby(
            ["GEO_NAME", "LATITUDE", "LONGITUDE", "H3_RES_3", "H3_RES_4", "H3_RES_5",
             "H3_RES_6", "H3_RES_7"], as_index=False)["CLIENT_COUNT"].sum()
    else:
        data = get_map_data(df_db_schema, f["q_to"], map_level, f["benefit_groups"] or ["Jobseeker Support"])
        subtitle = "Selected benefit groups, %s" % f["q_to"]

    if data.empty:
        st.warning("No mappable data for this selection.")
        return

    render_h3_map(data, resolution, jump, subtitle)

    st.markdown("---")
    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Area detail")
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(data.drop(columns=[c for c in data.columns if c.startswith("H3_")]),
                                    "MSD recipients by area — %s" % subtitle, "Area detail"),
            file_name="msd_area_detail_%s.xlsx" % _safe_filename(subtitle),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_map_detail_xlsx", type="primary")
    st.dataframe(
        data.drop(columns=[c for c in data.columns if c.startswith("H3_")]),
        use_container_width=True, hide_index=True)


def render_tab_housing(f):
    """Housing — emergency housing and the social housing register.

      two columns: grants and households granted (lines) | amount granted (bars)
      ---
      Housing Register and Transfer Register applications (lines)
      ---
      households by weeks-in-emergency-housing band (heatmap)
      ---
      territorial authority detail for one quarter, with an Excel export

    Each block is skipped silently when its series is absent, so a trimmed mart
    renders fewer blocks rather than failing.
    """
    st.header("Emergency housing and the social housing register")

    eh = get_emergency_housing_trend(df_db_schema)
    if not eh.empty:
        render_emergency_housing_charts(eh)

    st.markdown("---")
    reg = get_register_trend(df_db_schema)
    if not reg.empty:
        render_register_chart(reg)

    st.markdown("---")
    dur = get_eh_duration_matrix(df_db_schema)
    if not dur.empty:
        render_duration_heatmap(dur)

    st.markdown("---")
    render_eh_ta_detail(f)


def render_tab_hardship(f):
    """Hardship — supplementary support and hardship assistance.

      two columns: point-in-time recipients | assistance granted during the month
      ---
      3:2 columns: food grants nationally (bars) | latest month by region (bars)
      ---
      assistance paid by Work and Income region (stacked area), then a detail table
    """
    st.header("Supplementary support and hardship assistance")

    hard = get_hardship_trend(df_db_schema)
    if not hard.empty:
        render_hardship_charts(hard)

    st.markdown("---")
    food = get_food_grants(df_db_schema)
    if not food.empty:
        render_food_grant_charts(food)

    st.markdown("---")
    render_supplementary_detail(f)


def render_tab_studylink(f):
    """StudyLink — Student Allowance and Student Loans.

      2:2:3 columns: reporting window | tertiary measure for the right axis | caption
      combination chart: dollars paid as bars on the left axis, StudyLink
        recipients and total tertiary participation as lines on the right
      reach KPIs, then a recipients and a dollars tile per product
      ---
      two columns: recipients by provider type, one chart per product
      share of students at each provider type drawing support
      ---
      provider detail table with an Excel export

    The window picker exists because StudyLink republishes its whole history
    every quarter, each year truncated to the release quarter. Mixing windows
    splices a three-month year onto full ones, which is what made total support
    appear to rise while every component fell.
    """
    st.header("Student Allowance and Student Loans")

    windows = get_studylink_windows(df_db_schema).WINDOW_MONTHS.tolist()
    labels = {3: "January to March", 6: "January to June",
              9: "January to September", 12: "January to December (full year)"}
    c1, c1b, c2 = st.columns([2, 2, 3])
    with c1:
        window = st.selectbox(
            "Reporting window", options=sorted(windows, reverse=True),
            format_func=lambda w: labels.get(int(w), "%d months" % w),
            index=0, key="sl_window")
    with c1b:
        measure = st.selectbox(
            "Tertiary measure (right axis)", options=["ENROLMENTS", "EFTS"],
            format_func=lambda m: {"ENROLMENTS": "Students enrolled (headcount)",
                                   "EFTS": "Equivalent full-time students"}[m],
            key="sl_measure")
    with c2:
        st.caption(
            "StudyLink republishes its whole history every quarter with each year "
            "cut to the release quarter. Windows are never mixed here: pick one and "
            "every year on the charts covers the same months."
        )

    annual = get_studylink_annual(df_db_schema, int(window))
    if annual.empty:
        st.warning("No StudyLink data for this window.")
        return

    render_studylink_headline(annual, int(window), measure)
    st.markdown("---")
    render_studylink_provider_chart(int(window), measure)
    st.markdown("---")
    render_studylink_detail(annual, int(window))


def render_tab_pipeline(f):
    """Pipeline — where every figure came from.

      four KPI tiles: files downloaded, mart fact rows, publication coverage,
        missing releases
      ---
      coverage heatmap, series by period: green published, red missing
      ---
      staging coverage by fact, with an Excel export
      expander listing raw worksheets deliberately not staged, and why
      ---
      full download manifest with an Excel export
      ---
      provenance notes, also offered as copyable markdown

    This tab is the evidence for the banner at the top of the page. Every claim
    it makes is a query against the pipeline's own metadata, not prose.
    """
    st.header("Pipeline, coverage and provenance")

    cat = get_fact_catalog(df_db_schema)
    gaps = get_pipeline_gaps(df_db_schema)
    man = get_pipeline_manifest(df_db_schema)
    cov = get_pipeline_coverage(df_db_schema)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Source files downloaded", "{:,}".format(len(man)))
    c2.metric("Mart fact rows", "{:,}".format(int(cat.ROWS.sum())))
    present = int(gaps.IS_PRESENT.sum()) if not gaps.empty else 0
    c3.metric("Publication coverage",
              "{:.1f}%".format(100 * present / len(gaps)) if len(gaps) else "n/a")
    c4.metric("Missing releases", "{:,}".format(len(gaps) - present))

    st.markdown("---")
    render_coverage_matrix(gaps)

    st.markdown("---")
    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Staging coverage by fact")
    with dl:
        st.download_button("📥 Excel", data=build_styled_excel(cov, "Staging coverage by fact", "Coverage"),
                           file_name="msd_staging_coverage.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_cov_xlsx", type="primary")
    st.dataframe(cov, use_container_width=True, hide_index=True)

    with st.expander("Source worksheets not staged, and why"):
        skipped = get_skipped_sources(df_db_schema)
        st.caption(
            "Contents and notes worksheets carry no data. The remainder are pre-2014 "
            "layouts and the weekly, COVID wage-subsidy and Treasury forecast series, "
            "which are landed in RAW but out of scope for these tabs."
        )
        st.dataframe(skipped, use_container_width=True, hide_index=True)

    st.markdown("---")
    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Download manifest")
    with dl:
        st.download_button("📥 Excel",
                           data=build_styled_excel(man.head(50000), "MSD download manifest", "Manifest"),
                           file_name="msd_download_manifest.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_man_xlsx", type="primary")
    st.dataframe(man, use_container_width=True, hide_index=True, height=320)

    st.markdown("---")
    render_provenance_notes(gaps, man, cat)



def render_tab_retirement(f):
    """Retirement income — New Zealand Superannuation and the Veteran's Pension.

      caption noting recipients are MSD's and the cost is Treasury's
      recipients (left axis, lines) against NZ Super cost (right axis, solid
        outturn and dotted forecast), then four KPI tiles
      ---
      up to three stacked bar charts of recipient characteristics for a quarter
      ---
      hexagon map by territorial authority, with a detail table and export

    The two pensions stay separate on every visual: they share row labels in the
    source worksheets, and were collapsing into one another before the mart
    split them by section.
    """
    st.header("Guaranteed retirement income")
    st.caption(
        "New Zealand Superannuation and the Veteran's Pension: the universal, "
        "non-means-tested payments from age 65. Recipient counts are MSD's; the "
        "cost of them is Treasury's, because MSD publishes no expenditure at all."
    )

    summary = get_retirement_summary(df_db_schema)
    if summary.empty:
        st.warning("No retirement income data available.")
        return

    spend = get_expense_line(df_db_schema, ["New Zealand Superannuation", "Veteran's Pension"])
    render_retirement_headline(summary, spend)
    st.markdown("---")
    render_retirement_characteristics(f, summary)
    st.markdown("---")
    render_retirement_map(f)


def render_tab_all_assistance(f):
    """All assistance — every programme's spend on one axis.

      caption naming which source supplies spending and which supplies students
      stacked area of the ten largest programmes, with a dotted rule marking
        where outturn ends and forecast begins, then four KPI tiles
      ---
      people supported by programme (lines), with the standing warning not to add them
      ---
      full programme detail table with an Excel export
    """
    st.header("All government assistance")
    st.caption(
        "Main benefits, retirement income and student support on one axis. "
        "Spending is Treasury's core Crown expense tables; student support is "
        "MSD's own published totals. Forecast years are shown but never mixed "
        "with outturns."
    )

    assistance = get_all_assistance(df_db_schema)
    if assistance.empty:
        st.warning("No assistance data available.")
        return

    render_assistance_spend(assistance)
    st.markdown("---")
    render_assistance_people()
    st.markdown("---")
    render_assistance_detail(assistance)

def render_tab_build_notes(f):
    """Build Notes — the build write-up, rendered from its own markdown file.

      caption naming the document and stating it is loaded, not embedded
      3:1 header row: heading | download the original markdown
      the document, verbatim

    The text is read at run time rather than held in this module, so the tab
    always shows the current write-up and there is never a second copy of it to
    fall out of date. Nothing here interprets the document: it is displayed as
    the markdown it already is.
    """
    st.header("How this was built")
    st.caption(
        "The build write-up, loaded from %s and rendered unchanged. It records how "
        "the staging layer, the mart and this application were designed backwards "
        "from the question each tab has to answer." % BUILD_NOTES_DOC
    )

    doc = get_reference_doc(BUILD_NOTES_DOC)
    if doc is None:
        st.info(
            "%s is not on the reference path for this deployment. Point "
            "MSD_REFERENCE_DIR at the folder holding it." % BUILD_NOTES_DOC
        )
        return

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📄 %s" % BUILD_NOTES_DOC)
    with dl:
        st.download_button(
            "📥 Markdown", data=doc.encode("utf-8"), file_name=BUILD_NOTES_DOC,
            mime="text/markdown", key="dl_build_notes_md", type="primary")

    st.markdown("---")
    st.markdown(doc)


def render_header():
    """Hazard-striped provenance banner shown above every tab.

    Drawn as raw HTML rather than st.warning so the stripes survive both
    Streamlit themes: a 3px black frame, a 14px yellow-and-black 45-degree
    stripe band top and bottom, and a solid #FFD100 panel between them.

    The application is built from public releases but is not published by, nor
    endorsed by, the agencies that produced them, so that is stated before any
    figure is shown rather than buried in a footnote.
    """
    st.html(
        """
        <div style="border:3px solid #111; border-radius:6px; overflow:hidden;
                    margin:0 0 14px 0; font-family:sans-serif;">
          <div style="height:14px; background:repeating-linear-gradient(
                        45deg, #FFD100 0 14px, #111 14px 28px);"></div>
          <div style="background:#FFD100; color:#111; padding:12px 16px;">
            <div style="font-weight:800; font-size:15px; letter-spacing:.02em;">
              &#9888;&#65039; BUILT FROM NEW ZEALAND GOVERNMENT DATA &mdash;
              NOT AN OFFICIAL GOVERNMENT PRODUCT
            </div>
            <div style="font-size:13.5px; line-height:1.5; margin-top:6px;">
              Figures are reproduced from public releases by the
              <b>Ministry of Social Development</b> (including <b>StudyLink</b>),
              the <b>Ministry of Education</b> and <b>the Treasury</b>.
              This application is produced independently by Celnic Consulting and
              <b>does not represent the views, policy or official statistics of those
              departments</b>. Every original source file, with its download date and
              checksum, is listed in the <b>&#9881;&#65039; Pipeline</b> tab.
            </div>
          </div>
          <div style="height:14px; background:repeating-linear-gradient(
                        45deg, #FFD100 0 14px, #111 14px 28px);"></div>
        </div>
        """
    )


def render_attribution():
    """Collapsed source-and-licence note closing every page.

    The hazard banner above says who published the figures and that this is not
    an official product. This says the other half: under what terms the data is
    reused, that it has been modified to get here, and where the full per-dataset
    record lives. Collapsed, because it belongs after the reading rather than
    before it.
    """
    with st.expander("Data sources & attribution"):
        st.markdown(
            "Built on open data from the **Ministry of Social Development**, the "
            "**Ministry of Education** and **the Treasury** — modified: reshaped "
            "out of their worksheet layouts into a conformed mart. Licences were "
            "verified at source on 2026-08-30: **CC BY 4.0** for 24 of the 25 "
            "datasets, **CC BY 3.0 NZ** for the legacy 2017 Child, Youth and Family "
            "series. Ten of the 25 rest on an agency-wide statement rather than a "
            "dataset-specific one and stay flagged for confirmation. "
            "Demonstration of method, not published statistics. Full provenance: "
            "[ATTRIBUTION.md](https://github.com/celnicconsulting/msd-social-support-explorer/blob/main/ATTRIBUTION.md)."
        )


# ====================VISUALISATION====================
# One function per chart or table. The docstring is the specification of that
# visual: the chart type, what sits on each axis, the colour scheme, and the
# reason for the choice wherever it is not self-evident.
def render_overview_kpis(trend):
    """Headline counts with quarter-on-quarter and same-quarter year-on-year change.

    Four equal tiles: latest recipients, change on the previous quarter, change
    on the same quarter a year earlier, and how many quarters are on the charts.
    Both change tiles use inverse delta colouring, because more people on a
    benefit is not an improvement.

    Year-on-year compares the same quarter twelve months apart: the series is
    strongly seasonal and consecutive quarters are not comparable.
    """
    totals = trend.groupby(["PERIOD", "PERIOD_SORT"], as_index=False)["CLIENT_COUNT"].sum()
    totals = totals.sort_values("PERIOD_SORT")
    latest = totals.iloc[-1]
    prev = totals.iloc[-2] if len(totals) > 1 else None
    yoy_row = totals[totals.PERIOD_SORT == latest.PERIOD_SORT - 100]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Recipients, %s" % latest.PERIOD, "{:,.0f}".format(latest.CLIENT_COUNT))
    if prev is not None:
        d = latest.CLIENT_COUNT - prev.CLIENT_COUNT
        c2.metric("Change on %s" % prev.PERIOD, "{:+,.0f}".format(d),
                  delta="{:+.1f}%".format(100 * d / prev.CLIENT_COUNT), delta_color="inverse")
    if not yoy_row.empty:
        base = yoy_row.iloc[0]
        d = latest.CLIENT_COUNT - base.CLIENT_COUNT
        c3.metric("Change on %s (same quarter)" % base.PERIOD, "{:+,.0f}".format(d),
                  delta="{:+.1f}%".format(100 * d / base.CLIENT_COUNT), delta_color="inverse")
    else:
        c3.metric("Year-on-year", "n/a", help="Needs the same quarter twelve months earlier.")
    c4.metric("Quarters shown", "{:,}".format(len(totals)))


def render_benefit_area_chart(trend):
    """Stacked area: quarterly recipients by benefit group.

    Benefit colours are fixed in BENEFIT_COLOURS so a group keeps its colour on
    every tab. Unified hover, legend below the plot, 420px tall to sit level
    with the line chart beside it.
    """
    st.markdown("##### Recipients by benefit group")
    fig = px.area(trend.sort_values("PERIOD_SORT"), x="PERIOD", y="CLIENT_COUNT",
                  color="BENEFIT_GROUP", color_discrete_map=BENEFIT_COLOURS,
                  labels={"PERIOD": "", "CLIENT_COUNT": "Recipients", "BENEFIT_GROUP": ""})
    fig.update_layout(height=420, hovermode="x unified", legend_orientation="h",
                      legend_y=-0.2, margin=dict(t=10, b=10, l=0, r=0))
    fig.update_yaxes(tickformat=",")
    st.plotly_chart(fig, use_container_width=True)


def render_monthly_headline(f):
    """Monthly lines for the four headline series, beside the area chart.

    Restricted to All main benefits and the three largest so the panel stays
    readable at half width; the full monthly detail lives in the mart, not here.
    """
    st.markdown("##### Monthly headline series")
    monthly = get_monthly_trend(df_db_schema, 0, f["m_focus_sort"])
    if monthly.empty:
        st.info("No monthly data available.")
        return
    keep = ["All main benefits", "Jobseeker Support (JS)", "Sole Parent Support",
            "Supported Living Payment"]
    m = monthly[monthly.BENEFIT.isin(keep)]
    fig = px.line(m.sort_values("PERIOD_SORT"), x="PERIOD", y="CLIENT_COUNT", color="BENEFIT",
                  color_discrete_sequence=PALETTE,
                  labels={"PERIOD": "", "CLIENT_COUNT": "Recipients", "BENEFIT": ""})
    fig.update_layout(height=420, hovermode="x unified", legend_orientation="h",
                      legend_y=-0.25, margin=dict(t=10, b=10, l=0, r=0))
    fig.update_yaxes(tickformat=",")
    st.plotly_chart(fig, use_container_width=True)


def render_characteristic_charts(f):
    """Three horizontal bar charts side by side: age, gender, ethnic group.

    Sorted ascending so the longest bar sits at the top, and one colour
    throughout because the categories are not comparable with each other.

    The ethnicity chart captions its own basis and, on total response, states
    that the bars do not sum to the total.
    """
    st.markdown("##### Recipient characteristics, %s" % f["q_to"])
    if not f["geos"] or not f["benefit_groups"]:
        st.info("Select at least one area and benefit group.")
        return

    cols = st.columns(3)
    for col, char in zip(cols, ["Age group", "Gender", "Ethnic group"]):
        with col:
            d = get_characteristic_breakdown(df_db_schema, f["q_to"], f["geo_level"],
                                             f["geos"], f["benefit_groups"], char)
            if d.empty:
                st.info("No %s data for this selection." % char.lower())
                continue
            agg = d.groupby("CHARACTERISTIC_VALUE", as_index=False)["CLIENT_COUNT"].sum()
            agg = agg.sort_values("CLIENT_COUNT", ascending=True)
            fig = px.bar(agg, x="CLIENT_COUNT", y="CHARACTERISTIC_VALUE", orientation="h",
                         color_discrete_sequence=[PALETTE[0]],
                         labels={"CLIENT_COUNT": "Recipients", "CHARACTERISTIC_VALUE": ""})
            fig.update_layout(height=300, margin=dict(t=30, b=10, l=0, r=0),
                              title=dict(text=char, font=dict(size=14)))
            fig.update_xaxes(tickformat=",")
            st.plotly_chart(fig, use_container_width=True)
            if char == "Ethnic group":
                basis = d.ETHNICITY_BASIS.iloc[0]
                if basis == "TOTAL_RESPONSE":
                    st.caption(
                        "⚠️ Total response basis: a person appears in every ethnic group "
                        "they identify with, so these bars do not sum to the total."
                    )
                else:
                    st.caption("Prioritised ethnicity basis (before the December 2021 quarter).")


def render_snapshot_table(f):
    """Area by benefit group pivot for the closing quarter, with a Total column.

    Uses the standard table header: a 3:1 column split that puts the Excel
    download button to the right of the heading.
    """
    snap = get_latest_snapshot(df_db_schema, f["q_to"], f["geo_level"], f["benefit_groups"])
    if snap.empty:
        return
    pivot = snap.pivot_table(index="GEOGRAPHY", columns="BENEFIT_GROUP",
                             values="CLIENT_COUNT", aggfunc="sum").fillna(0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False).reset_index()

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Snapshot by area — %s" % f["q_to"])
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(pivot, "MSD recipients by area — %s" % f["q_to"], "Snapshot"),
            file_name="msd_snapshot_%s.xlsx" % _safe_filename(f["q_to"]),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_snapshot_xlsx", type="primary")
    st.dataframe(pivot, use_container_width=True, hide_index=True)


def render_h3_map(df, resolution, jump, subtitle):
    """H3 hexagon layer coloured red-to-yellow by client count.

    Shared by the Map and Retirement tabs, so it takes a plain dataframe with a
    CLIENT_COUNT column and pre-computed H3 cells rather than querying anything
    itself.

    Colour runs yellow (low) to red (high) across the range currently in view,
    and hexagons are extruded in proportion, so both channels carry the same
    number. The basemap is a CartoDB style served over https: pydeck's default
    mapbox:// style needs a token the Community Cloud deployment does not have.
    """
    col = "H3_RES_%d" % resolution
    d = df[df[col].notna()].copy()
    if d.empty:
        st.warning("No hexagons at this resolution.")
        return

    d = d.groupby([col], as_index=False).agg(
        CLIENT_COUNT=("CLIENT_COUNT", "sum"),
        AREAS=("GEO_NAME", lambda s: ", ".join(sorted(set(s))[:4])),
        LATITUDE=("LATITUDE", "mean"), LONGITUDE=("LONGITUDE", "mean"))
    d = d.rename(columns={col: "HEX"})

    lo, hi = d.CLIENT_COUNT.min(), d.CLIENT_COUNT.max()
    span = (hi - lo) or 1

    def colour(v):
        t = (v - lo) / span
        return [255, int(40 + 215 * (1 - t)), int(20 + 40 * (1 - t)), 190]

    d["COLOR"] = d.CLIENT_COUNT.apply(colour)
    d["ELEVATION"] = (d.CLIENT_COUNT / span * 40000).clip(lower=500)

    views = {"New Zealand": (-41.0, 173.5, 4.6), "Auckland": (-36.85, 174.76, 8.5),
             "Wellington": (-41.2865, 174.7762, 8.5), "Christchurch": (-43.53, 172.64, 8.5),
             "Hamilton": (-37.79, 175.28, 8.5), "Dunedin": (-45.87, 170.50, 8.5)}
    lat, lon, zoom = views[jump]

    layer = pdk.Layer(
        "H3HexagonLayer", d, get_hexagon="HEX", get_fill_color="COLOR",
        get_elevation="ELEVATION", elevation_scale=1, extruded=True,
        get_line_color=[255, 255, 255, 60], pickable=True, stroked=True, filled=True,
        line_width_min_pixels=1)

    st.caption(subtitle)
    st.pydeck_chart(pdk.Deck(
        layers=[layer],
        initial_view_state=pdk.ViewState(latitude=lat, longitude=lon, zoom=zoom, pitch=35, bearing=0),
        tooltip={"html": "<b>{AREAS}</b><br/>Recipients: {CLIENT_COUNT}",
                 "style": {"backgroundColor": "#1A3A5C", "color": "white"}},
        map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"))


def render_emergency_housing_charts(eh):
    """Two equal columns: grant and household counts as lines, dollars as bars.

    Counts and dollars get separate charts rather than a shared secondary axis,
    because they are different VALUE_KINDs and one axis invites reading one as
    the other.
    """
    st.markdown("##### Emergency Housing Special Needs Grants, monthly")
    c1, c2 = st.columns(2)
    with c1:
        d = eh[eh.METRIC.isin(["Number of EHGs granted", "Number of households granted EHGs"])]
        fig = px.line(d.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE", color="METRIC",
                      color_discrete_sequence=PALETTE, labels={"PERIOD": "", "VALUE": "", "METRIC": ""})
        fig.update_layout(height=340, hovermode="x unified", legend_orientation="h",
                          legend_y=-0.3, margin=dict(t=10, b=10, l=0, r=0))
        fig.update_yaxes(tickformat=",")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        d = eh[(eh.METRIC == "Total amount granted")]
        fig = px.bar(d.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE",
                     color_discrete_sequence=[PALETTE[1]],
                     labels={"PERIOD": "", "VALUE": "Amount granted (NZ$)"})
        fig.update_layout(height=340, margin=dict(t=10, b=10, l=0, r=0), showlegend=False)
        fig.update_yaxes(tickprefix="$", tickformat=",")
        st.plotly_chart(fig, use_container_width=True)


def render_register_chart(reg):
    """Two lines: Housing Register against Transfer Register applications."""
    st.markdown("##### Housing Register and Transfer Register")
    d = reg.copy()
    fig = px.line(d.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE", color="REGISTER",
                  color_discrete_sequence=[PALETTE[0], PALETTE[1]],
                  labels={"PERIOD": "", "VALUE": "Applications", "REGISTER": ""})
    fig.update_layout(height=360, hovermode="x unified", legend_orientation="h",
                      legend_y=-0.25, margin=dict(t=10, b=10, l=0, r=0))
    fig.update_yaxes(tickformat=",")
    st.plotly_chart(fig, use_container_width=True)


def render_duration_heatmap(dur):
    """Duration band by quarter heatmap, pale yellow (few) to red (many households).

    Columns are ordered by PERIOD_SORT rather than alphabetically, because the
    period labels do not sort as dates.
    """
    st.markdown("##### Households in emergency housing by duration band")
    p = dur.pivot_table(index="DURATION_BAND", columns="PERIOD", values="HOUSEHOLDS", aggfunc="sum")
    order = sorted(p.columns, key=lambda c: dur[dur.PERIOD == c].PERIOD_SORT.iloc[0])
    p = p[order]
    fig = go.Figure(go.Heatmap(z=p.values, x=p.columns, y=p.index,
                               colorscale=[[0, "#FFF3B0"], [0.5, "#F6AE2D"], [1, "#C33C54"]],
                               hovertemplate="%{y}<br>%{x}: %{z:,.0f} households<extra></extra>"))
    fig.update_layout(height=320, margin=dict(t=10, b=10, l=0, r=0))
    st.plotly_chart(fig, use_container_width=True)


def render_eh_ta_detail(f):
    """Territorial authority table for one emergency housing quarter.

    The picker starts at 2022Q1, which is when MSD began publishing the TA
    breakdown. Blank cells are captioned as suppressions so no reader treats
    them as zero.
    """
    quarters = sorted(
        get_periods(df_db_schema, "QUARTER").PERIOD.tolist())
    eh_quarters = [q for q in quarters if q >= "2022Q1"]
    period = st.selectbox("Emergency housing quarter", options=eh_quarters[::-1], index=0,
                          key="eh_ta_period")
    d = get_eh_by_ta(df_db_schema, period)
    if d.empty:
        st.info("No territorial authority data for %s." % period)
        return

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Emergency housing by territorial authority — %s" % period)
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(d, "Emergency housing by TA — %s" % period, "EH by TA"),
            file_name="msd_emergency_housing_ta_%s.xlsx" % _safe_filename(period),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ehta_xlsx", type="primary")
    st.caption(
        "Blank values are MSD confidentiality suppressions, flagged in IS_SUPPRESSED. "
        "They are not zero."
    )
    st.dataframe(d, use_container_width=True, hide_index=True, height=380)


def render_hardship_charts(hard):
    """Two equal columns of lines: point-in-time recipients | recipients during the month.

    The split matters. Point-in-time counts compare across months; the
    within-month counts are flows and do not.
    """
    st.markdown("##### Supplementary support and hardship assistance, monthly")
    c1, c2 = st.columns(2)
    point = hard[hard.SECTION.str.contains("Point-in-time", case=False, na=False)
                 | hard.ASSISTANCE_TYPE.isin(["Accommodation Supplement", "Disability Allowance"])]
    cum = hard[hard.ASSISTANCE_TYPE.str.contains("Special Needs|Advances", case=False, na=False)]
    with c1:
        d = point[point.VALUE_KIND == "COUNT"]
        fig = px.line(d.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE", color="ASSISTANCE_TYPE",
                      color_discrete_sequence=PALETTE,
                      labels={"PERIOD": "", "VALUE": "Recipients", "ASSISTANCE_TYPE": ""})
        fig.update_layout(height=360, hovermode="x unified", legend_orientation="h",
                          legend_y=-0.35, margin=dict(t=30, b=10, l=0, r=0),
                          title=dict(text="Point-in-time supplementary support", font=dict(size=14)))
        fig.update_yaxes(tickformat=",")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        d = cum[cum.VALUE_KIND == "COUNT"]
        fig = px.line(d.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE", color="ASSISTANCE_TYPE",
                      color_discrete_sequence=PALETTE[1:],
                      labels={"PERIOD": "", "VALUE": "Recipients in month", "ASSISTANCE_TYPE": ""})
        fig.update_layout(height=360, hovermode="x unified", legend_orientation="h",
                          legend_y=-0.35, margin=dict(t=30, b=10, l=0, r=0),
                          title=dict(text="Hardship assistance during the month", font=dict(size=14)))
        fig.update_yaxes(tickformat=",")
        st.plotly_chart(fig, use_container_width=True)


def render_food_grant_charts(food):
    """3:2 columns: national food grants per month (bars) | latest month by region (bars).

    The regional chart keeps the top twelve regions so the labels stay legible.
    """
    st.markdown("##### Special Needs Grants for food, by regional council")
    counts = food[food.VALUE_KIND == "COUNT"]
    if counts.empty:
        st.info("No food grant counts available.")
        return
    c1, c2 = st.columns([3, 2])
    with c1:
        nat = counts.groupby(["PERIOD", "PERIOD_SORT"], as_index=False)["VALUE"].sum()
        fig = px.bar(nat.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE",
                     color_discrete_sequence=[PALETTE[1]],
                     labels={"PERIOD": "", "VALUE": "Food grants"})
        fig.update_layout(height=340, margin=dict(t=30, b=10, l=0, r=0), showlegend=False,
                          title=dict(text="All regions", font=dict(size=14)))
        fig.update_yaxes(tickformat=",")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        latest = counts[counts.PERIOD_SORT == counts.PERIOD_SORT.max()]
        agg = latest.groupby("REGIONAL_COUNCIL", as_index=False)["VALUE"].sum()
        agg = agg.sort_values("VALUE", ascending=True).tail(12)
        fig = px.bar(agg, x="VALUE", y="REGIONAL_COUNCIL", orientation="h",
                     color_discrete_sequence=[PALETTE[0]],
                     labels={"VALUE": "Food grants", "REGIONAL_COUNCIL": ""})
        fig.update_layout(height=340, margin=dict(t=30, b=10, l=0, r=0),
                          title=dict(text="Latest month by region", font=dict(size=14)))
        fig.update_xaxes(tickformat=",")
        st.plotly_chart(fig, use_container_width=True)


def render_supplementary_detail(f):
    """Stacked area of assistance paid by Work and Income region, then a detail table.

    Only the AMOUNT rows are charted. The count rows share the fact and would
    land on the same axis at a completely different scale.
    """
    d = get_supplementary_region(df_db_schema, f["q_from_sort"], f["q_to_sort"])
    if d.empty:
        st.info("No regional supplementary assistance data in this period range.")
        return
    amounts = d[d.VALUE_KIND == "AMOUNT"]
    if not amounts.empty:
        st.markdown("##### Supplementary and hardship assistance paid, by Work and Income region")
        agg = amounts.groupby(["PERIOD", "PERIOD_SORT", "WI_REGION"], as_index=False)["VALUE"].sum()
        fig = px.area(agg.sort_values("PERIOD_SORT"), x="PERIOD", y="VALUE", color="WI_REGION",
                      color_discrete_sequence=px.colors.qualitative.Safe,
                      labels={"PERIOD": "", "VALUE": "Amount paid (NZ$)", "WI_REGION": ""})
        fig.update_layout(height=380, hovermode="x unified", legend_orientation="h",
                          legend_y=-0.3, margin=dict(t=10, b=10, l=0, r=0))
        fig.update_yaxes(tickprefix="$", tickformat=",")
        st.plotly_chart(fig, use_container_width=True)

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Regional supplementary assistance detail")
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(d.head(50000), "Supplementary assistance by region", "Supplementary"),
            file_name="msd_supplementary_region.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_supp_xlsx", type="primary")
    st.dataframe(d, use_container_width=True, hide_index=True, height=340)


def render_studylink_headline(annual, window, measure):
    """Support paid on the left axis, students on the right.

    Bars are dollars paid per product on the left axis. Solid lines are
    StudyLink recipients and a dotted navy line is total tertiary participation,
    both on the right axis. Below the chart sit the reach KPIs, then a
    recipients tile and a dollars tile for each product.

    The right axis carries StudyLink recipients and, when the Education Counts
    extract has been loaded, all tertiary enrolments beside them, so a fall in
    students receiving support can be read against total participation. This is
    the pairing that showed university attendance rising while loan uptake fell.
    """
    st.markdown("##### Support paid and students supported")

    enrol = get_tertiary_total(df_db_schema, measure)
    fig = go.Figure()

    for i, product in enumerate(sorted(annual.PRODUCT.unique())):
        d = annual[annual.PRODUCT == product].sort_values("PERIOD_SORT")
        fig.add_trace(go.Bar(
            x=d.PERIOD, y=d.AMOUNT_NZD, name="%s paid" % product,
            marker_color=PALETTE[i], opacity=0.85,
            hovertemplate="%{x}<br>%{fullData.name}: $%{y:,.0f}<extra></extra>"))

    for i, product in enumerate(sorted(annual.PRODUCT.unique())):
        d = annual[annual.PRODUCT == product].sort_values("PERIOD_SORT")
        fig.add_trace(go.Scatter(
            x=d.PERIOD, y=d.RECIPIENTS, name="%s recipients" % product,
            yaxis="y2", mode="lines+markers", line=dict(color=PALETTE[i], width=3),
            marker=dict(size=6),
            hovertemplate="%{x}<br>%{fullData.name}: %{y:,.0f}<extra></extra>"))

    label = ("All tertiary students (MoE)" if measure == "ENROLMENTS"
             else "All tertiary EFTS (MoE)")
    if not enrol.empty:
        tot = enrol.sort_values("PERIOD_SORT")
        fig.add_trace(go.Scatter(
            x=tot.PERIOD, y=tot.VALUE, name=label,
            yaxis="y2", mode="lines", line=dict(color="#1A3A5C", width=3, dash="dot"),
            hovertemplate="%{x}<br>" + label + ": %{y:,.0f}<extra></extra>"))

    fig.update_layout(
        height=460, barmode="group", hovermode="x unified",
        legend=dict(orientation="h", y=-0.22),
        margin=dict(t=10, b=10, l=0, r=0),
        yaxis=dict(title="Paid (NZ$)", tickprefix="$", tickformat=","),
        yaxis2=dict(title="Students", overlaying="y", side="right",
                    tickformat=",", showgrid=False))
    st.plotly_chart(fig, use_container_width=True)

    if enrol.empty:
        st.caption(
            "Ministry of Education tertiary participation is not loaded, so the right "
            "axis shows StudyLink recipients only."
        )
    else:
        st.caption(
            "Right axis: StudyLink recipients against all tertiary participation "
            "(Education Counts, provider-based, 2016-2025). StudyLink counts people who "
            "received support during the window; the Ministry counts everyone enrolled "
            "at any time in the year, so the gap between the lines is students studying "
            "without a loan or allowance."
        )
        render_support_share(annual, enrol, measure)

    latest = annual.sort_values("PERIOD_SORT").PERIOD.iloc[-1]
    cols = st.columns(len(annual.PRODUCT.unique()) * 2)
    for i, product in enumerate(sorted(annual.PRODUCT.unique())):
        d = annual[annual.PRODUCT == product].sort_values("PERIOD_SORT")
        last, first = d.iloc[-1], d.iloc[0]
        cols[i * 2].metric("%s recipients, %s" % (product, latest),
                           "{:,.0f}".format(last.RECIPIENTS) if pd.notna(last.RECIPIENTS) else "n/a",
                           delta="{:+,.0f} since {}".format(last.RECIPIENTS - first.RECIPIENTS, first.PERIOD)
                           if pd.notna(last.RECIPIENTS) and pd.notna(first.RECIPIENTS) else None)
        cols[i * 2 + 1].metric("%s paid, %s" % (product, latest),
                               "${:,.0f}".format(last.AMOUNT_NZD) if pd.notna(last.AMOUNT_NZD) else "n/a")


def render_support_share(annual, enrol, measure):
    """How much of the tertiary population StudyLink reaches.

    Three tiles: all tertiary students in the latest year both sources cover,
    then each product's recipients as a percentage of them. The share, not the
    count, is what answers whether fewer students are getting support or there
    are simply fewer students.
    """
    tot = enrol.set_index("PERIOD").VALUE
    years = [p for p in annual.PERIOD.unique() if p in tot.index]
    if not years:
        return
    y = max(years)
    a = annual[annual.PERIOD == y]
    cols = st.columns(3)
    cols[0].metric("All tertiary students, %s" % y, "{:,.0f}".format(tot[y]))
    for i, product in enumerate(sorted(a.PRODUCT.unique())):
        r = a[a.PRODUCT == product].RECIPIENTS.iloc[0]
        if pd.notna(r):
            cols[i + 1].metric("%s reach, %s" % (product, y),
                               "{:.1f}%".format(100 * r / tot[y]),
                               help="StudyLink recipients as a share of all tertiary students.")


def render_studylink_provider_chart(window, measure):
    """Recipients by provider type, one chart per product, then the support share.

    Two equal columns (Student Allowance | Student Loans), then a full-width
    percentage chart where colour is provider type and dash pattern is product.
    """
    st.markdown("##### Students supported, by education provider type")
    prov = get_studylink_provider(df_db_schema, window)
    if prov.empty:
        st.info("No provider breakdown for this window.")
        return

    # Only the single-provider categories are comparable with the Ministry's
    # subsectors; MSD also publishes combination buckets for students enrolled at
    # more than one provider type, which belong to neither.
    d = prov[prov.PROVIDER_TYPE_STD.notna()].copy()
    d = d.rename(columns={"PROVIDER_TYPE_STD": "PROVIDER_TYPE_LABEL"})
    d = d.groupby(["PERIOD", "PERIOD_SORT", "PRODUCT", "PROVIDER_TYPE_LABEL"],
                  as_index=False)["RECIPIENTS"].sum()
    d["PROVIDER_TYPE"] = d.PROVIDER_TYPE_LABEL

    c1, c2 = st.columns(2)
    for col, product in zip((c1, c2), sorted(d.PRODUCT.unique())):
        with col:
            sub = d[d.PRODUCT == product].sort_values("PERIOD_SORT")
            fig = px.line(sub, x="PERIOD", y="RECIPIENTS", color="PROVIDER_TYPE",
                          color_discrete_sequence=px.colors.qualitative.Safe,
                          labels={"PERIOD": "", "RECIPIENTS": "Students", "PROVIDER_TYPE": ""})
            fig.update_layout(height=360, hovermode="x unified", legend_orientation="h",
                              legend_y=-0.3, margin=dict(t=30, b=10, l=0, r=0),
                              title=dict(text=product, font=dict(size=14)))
            fig.update_yaxes(tickformat=",")
            st.plotly_chart(fig, use_container_width=True)

    moe = get_tertiary_by_provider(df_db_schema, measure)
    if not moe.empty:
        st.markdown("##### Share of students at each provider type receiving support")
        share = d.merge(moe.rename(columns={"VALUE": "ALL_STUDENTS"}),
                        left_on=["PERIOD", "PROVIDER_TYPE_LABEL"],
                        right_on=["PERIOD", "PROVIDER_TYPE_STD"], how="inner")
        share = share[share.ALL_STUDENTS > 0].copy()
        share["SHARE_PCT"] = 100 * share.RECIPIENTS / share.ALL_STUDENTS
        fig = px.line(share.sort_values("PERIOD_SORT_x"), x="PERIOD", y="SHARE_PCT",
                      color="PROVIDER_TYPE", line_dash="PRODUCT",
                      color_discrete_sequence=px.colors.qualitative.Safe,
                      labels={"PERIOD": "", "SHARE_PCT": "% of students", "PROVIDER_TYPE": ""})
        fig.update_layout(height=380, hovermode="x unified", legend_orientation="h",
                          legend_y=-0.35, margin=dict(t=10, b=10, l=0, r=0))
        fig.update_yaxes(ticksuffix="%")
        st.plotly_chart(fig, use_container_width=True)
        st.caption(
            "StudyLink recipients as a percentage of all students enrolled at that "
            "provider type. A falling line means fewer of the students who are there "
            "are drawing support, rather than fewer students."
        )


def render_studylink_detail(annual, window):
    """Provider-level StudyLink table for the chosen window, with an Excel export."""
    prov = get_studylink_provider(df_db_schema, window)
    detail = prov.drop(columns=["PERIOD_SORT"]) if not prov.empty else annual
    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 StudyLink by provider type — %d-month window" % window)
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(detail.head(50000),
                                    "StudyLink by provider type — %d-month window" % window,
                                    "StudyLink"),
            file_name="msd_studylink_provider_%dm.xlsx" % window,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_sl_xlsx", type="primary")
    st.dataframe(detail, use_container_width=True, hide_index=True, height=340)


def render_coverage_matrix(gaps):
    """Series by period heatmap of publication coverage: green published, red missing.

    Height grows with the number of series so rows never compress into a band.
    Missing releases are also listed in an expander, because a red cell raises a
    question and the list answers it.
    """
    st.markdown("##### Publication coverage by series")
    if gaps.empty:
        st.info("No coverage register available.")
        return
    piv = gaps.pivot_table(index="FAMILY", columns="PERIOD", values="IS_PRESENT",
                           aggfunc="max")
    piv = piv.reindex(sorted(piv.columns), axis=1)
    z = piv.astype(float).values
    fig = go.Figure(go.Heatmap(
        z=z, x=list(piv.columns), y=list(piv.index),
        colorscale=[[0, "#C33C54"], [1, "#4C9F70"]], showscale=False,
        hovertemplate="%{y}<br>%{x}: %{z}<extra></extra>"))
    fig.update_layout(height=max(320, 18 * len(piv)), margin=dict(t=10, b=10, l=0, r=0))
    st.plotly_chart(fig, use_container_width=True)
    st.caption("Green: release published and loaded. Red: release missing at the detected cadence.")

    missing = gaps[~gaps.IS_PRESENT]
    if not missing.empty:
        with st.expander("Missing releases (%d)" % len(missing)):
            st.dataframe(missing[["SCHEMA_NAME", "FAMILY", "CADENCE_MONTHS", "PERIOD"]],
                         use_container_width=True, hide_index=True)


def render_provenance_notes(gaps, man, cat):
    """The written provenance statement, with every number filled from the mart.

    Nothing in the prose is typed: the counts, coverage and period span are
    interpolated from the metadata queries above, so the statement cannot drift
    from the pipeline. Offered again as copyable markdown so it can be pasted
    into a report.
    """
    present = int(gaps.IS_PRESENT.sum()) if not gaps.empty else 0
    missing = len(gaps) - present
    notes = """
### Provenance

All figures are published Ministry of Social Development statistics, downloaded from
msd.govt.nz. **No synthetic, modelled or interpolated values appear anywhere in this
application.**

- **{files:,} source files** ({mb:.0f} MB) downloaded from the MSD statistics section.
- **{rows:,} fact rows** across {facts} conformed fact tables.
- **{cov:.1f}% publication coverage** — {missing} missing releases out of {slots}.
- Earliest period **{first}**, latest **{last}**.

### Reading the numbers

- **Suppression is not zero.** MSD withholds low counts for confidentiality. Those cells
  are held as NULL with an `IS_SUPPRESSED` flag and are never charted as zero.
- **Ethnicity changed basis in the December 2021 quarter.** Before it, ethnicity was
  prioritised — one group per person. From it, ethnicity is total response — a person
  counts in every group they identify with, so ethnicity columns must not be summed.
  Every row carries `ETHNICITY_BASIS`.
- **Compare quarters year-on-year, not consecutively.** Benefit numbers are seasonal;
  the Overview tab's headline change compares the same quarter twelve months apart.
- **Counts are randomly rounded** by MSD, so components may not sum exactly to totals.
  Regional sums reconcile to national totals within 0.01%.
- **Map hexagons mark indicative centroids**, not area boundaries. MSD publishes counts
  by area name only.

### Where the 2025/26 gap went

The brief for this platform anticipated a publication outage from roughly December 2025
to February 2026 caused by MSD's Information Analysis Platform, and specified synthetic
infill for the missing months. **That gap no longer exists**: the Monthly Benefits Update
series is complete month by month from April 2020 to July 2026. MSD backfilled the outage
months, so no infill was built.
""".format(files=len(man), mb=man.FILE_SIZE.sum() / 1e6, rows=int(cat.ROWS.sum()),
           facts=len(cat), cov=100 * present / len(gaps) if len(gaps) else 0,
           missing=missing, slots=len(gaps),
           first=cat.MIN_PERIOD.min(), last=cat.MAX_PERIOD.max())
    st.markdown(notes)
    with st.expander("Copy this analysis as markdown"):
        st.code(notes, language="markdown")



def render_retirement_headline(summary, spend):
    """Recipients on the left axis, what they cost on the right.

    A solid line per pension, then Treasury's NZ Super cost as a solid line for
    outturn and a dotted one for forecast — never a single series, so no reader
    mistakes a projection for a result. Four tiles below: each pension's latest
    count, the latest actual cost and the furthest forecast.
    """
    st.markdown("##### Recipients and cost")
    fig = go.Figure()
    colours = {"New Zealand Superannuation": PALETTE[0], "Veteran's Pension": PALETTE[1]}

    for pension in sorted(summary.PENSION.unique()):
        d = summary[summary.PENSION == pension].sort_values("PERIOD_SORT")
        fig.add_trace(go.Scatter(
            x=d.PERIOD, y=d.RECIPIENTS, name=pension, mode="lines",
            line=dict(color=colours.get(pension, PALETTE[2]), width=3),
            hovertemplate="%{x}<br>%{fullData.name}: %{y:,.0f}<extra></extra>"))

    if not spend.empty:
        for basis, dash in (("Actual", "solid"), ("Forecast", "dot")):
            d = spend[(spend.EXPENSE_LINE == "New Zealand Superannuation")
                      & (spend.BASIS == basis)].sort_values("PERIOD_SORT")
            if d.empty:
                continue
            fig.add_trace(go.Scatter(
                x=d.PERIOD, y=d.AMOUNT_NZD_MILLION,
                name="NZ Super cost (%s)" % basis.lower(), yaxis="y2", mode="lines",
                line=dict(color="#1A3A5C", width=3, dash=dash),
                hovertemplate="%{x}<br>%{fullData.name}: $%{y:,.0f}m<extra></extra>"))

    fig.update_layout(
        height=440, hovermode="x unified", legend=dict(orientation="h", y=-0.22),
        margin=dict(t=10, b=10, l=0, r=0),
        yaxis=dict(title="Recipients", tickformat=","),
        yaxis2=dict(title="Cost (NZ$m)", overlaying="y", side="right",
                    tickprefix="$", tickformat=",", showgrid=False))
    st.plotly_chart(fig, use_container_width=True)

    latest = summary.sort_values("PERIOD_SORT").PERIOD.iloc[-1]
    cur = summary[summary.PERIOD == latest]
    cols = st.columns(4)
    for i, pension in enumerate(sorted(cur.PENSION.unique())):
        v = cur[cur.PENSION == pension].RECIPIENTS.sum()
        cols[i].metric("%s, %s" % (pension, latest), "{:,.0f}".format(v))
    if not spend.empty:
        act = spend[(spend.EXPENSE_LINE == "New Zealand Superannuation") & (spend.BASIS == "Actual")]
        fc = spend[(spend.EXPENSE_LINE == "New Zealand Superannuation") & (spend.BASIS == "Forecast")]
        if not act.empty:
            a = act.sort_values("PERIOD_SORT").iloc[-1]
            cols[2].metric("NZ Super cost, %s" % a.PERIOD, "${:,.0f}m".format(a.AMOUNT_NZD_MILLION))
        if not fc.empty:
            b = fc.sort_values("PERIOD_SORT").iloc[-1]
            cols[3].metric("Forecast, %s" % b.PERIOD, "${:,.0f}m".format(b.AMOUNT_NZD_MILLION),
                           help="Treasury Budget Economic and Fiscal Update projection.")


def render_retirement_characteristics(f, summary):
    """Up to three stacked bar charts of recipient characteristics for a chosen quarter.

    Which characteristics appear is read from the data rather than hard-coded,
    because MSD's published breakdowns change between releases. Bars stack by
    pension in the same colours as the headline chart above.
    """
    st.markdown("##### Recipient characteristics")
    periods = sorted(summary.PERIOD.unique())
    period = st.selectbox("Quarter", options=periods[::-1], index=0, key="ret_char_period")
    types = get_retirement_characteristic_types(df_db_schema, period).CHARACTERISTIC.tolist()
    types = [t for t in types if t.lower() not in {"total", "recipient characteristic"}][:3]
    if not types:
        st.info("No characteristic breakdown published for %s." % period)
        return
    cols = st.columns(len(types))
    for col, char in zip(cols, types):
        with col:
            d = get_retirement_characteristics(df_db_schema, period, char)
            if d.empty:
                continue
            order = (d.groupby("CHARACTERISTIC_VALUE")["RECIPIENTS"].sum()
                     .sort_values().index.tolist())
            fig = px.bar(d, x="RECIPIENTS", y="CHARACTERISTIC_VALUE", orientation="h",
                         color="PENSION", barmode="stack",
                         color_discrete_map={"New Zealand Superannuation": PALETTE[0],
                                             "Veteran's Pension": PALETTE[1]},
                         category_orders={"CHARACTERISTIC_VALUE": order},
                         labels={"RECIPIENTS": "Recipients", "CHARACTERISTIC_VALUE": "",
                                 "PENSION": ""})
            fig.update_layout(height=340, margin=dict(t=30, b=10, l=0, r=0),
                              legend=dict(orientation="h", y=-0.25),
                              title=dict(text=char, font=dict(size=14)))
            fig.update_xaxes(tickformat=",")
            st.plotly_chart(fig, use_container_width=True)


def render_retirement_map(f):
    """Hexagon map of pension recipients by territorial authority.

    Three controls: quarter, pension and hexagon resolution. Pension is a
    control rather than a colour because the two are on wildly different
    scales — the Veteran's Pension would be invisible beside NZ Super.
    """
    st.markdown("##### Where recipients live")
    periods = get_retirement_periods(df_db_schema).PERIOD.tolist()
    if not periods:
        st.info("No territorial authority breakdown available.")
        return
    c1, c2, c3 = st.columns(3)
    with c1:
        period = st.selectbox("Quarter", options=periods[::-1], index=0, key="ret_map_period")
    with c2:
        pension = st.selectbox("Pension", ["New Zealand Superannuation", "Veteran's Pension"],
                               key="ret_map_pension")
    with c3:
        res = st.select_slider("Hexagon resolution", options=[3, 4, 5, 6, 7], value=4,
                               key="ret_map_res")

    d = get_retirement_ta(df_db_schema, period, pension)
    if d.empty:
        st.info("No mappable data for this selection.")
        return
    d = d.rename(columns={"RECIPIENTS": "CLIENT_COUNT"})
    render_h3_map(d, res, "New Zealand", "%s recipients by territorial authority, %s"
                  % (pension, period))

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Recipients by territorial authority")
    with dl:
        out = d.drop(columns=[c for c in d.columns if c.startswith("H3_")])
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(out, "%s by TA — %s" % (pension, period), "Retirement"),
            file_name="msd_retirement_ta_%s.xlsx" % _safe_filename(period),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ret_ta_xlsx", type="primary")
    st.dataframe(d.drop(columns=[c for c in d.columns if c.startswith("H3_")]),
                 use_container_width=True, hide_index=True, height=320)


def render_assistance_spend(assistance):
    """Total assistance spending by programme, outturn then projection.

    Stacked area of the ten largest programmes, with a dotted vertical rule and
    an "outturn to forecast" annotation at the last actual year. Four tiles
    below: total for that year, largest programme, the furthest benefit forecast
    and how many programmes are counted.
    """
    st.markdown("##### What government assistance costs")

    top = (assistance.groupby("PROGRAMME")["AMOUNT_NZD_MILLION"].max()
           .sort_values(ascending=False).head(10).index.tolist())
    d = assistance[assistance.PROGRAMME.isin(top)].copy()

    fig = px.area(d.sort_values("PERIOD_SORT"), x="PERIOD", y="AMOUNT_NZD_MILLION",
                  color="PROGRAMME", color_discrete_sequence=px.colors.qualitative.Safe,
                  labels={"PERIOD": "", "AMOUNT_NZD_MILLION": "NZ$ million", "PROGRAMME": ""})
    last_actual = d[d.BASIS == "Actual"].PERIOD.max()
    if pd.notna(last_actual):
        fig.add_vline(x=last_actual, line_width=2, line_dash="dot", line_color="#1A3A5C")
        fig.add_annotation(x=last_actual, yref="paper", y=1.02, showarrow=False,
                           text="outturn ← | → forecast", font=dict(size=11, color="#1A3A5C"))
    fig.update_layout(height=460, hovermode="x unified", legend=dict(orientation="h", y=-0.25),
                      margin=dict(t=30, b=10, l=0, r=0))
    fig.update_yaxes(tickprefix="$", tickformat=",")
    st.plotly_chart(fig, use_container_width=True)

    # Treasury outturns stop a year earlier than MSD's student figures, so the
    # headline year is the latest one both sources cover; otherwise the total
    # silently becomes student support alone.
    act = assistance[assistance.BASIS == "Actual"]
    tsy = act[act.SPEND_SOURCE.str.startswith("Treasury")]
    if not tsy.empty:
        y = tsy.PERIOD.max()
        year_rows = act[act.PERIOD == y]
        tot = year_rows.AMOUNT_NZD_MILLION.sum()
        fc = assistance[(assistance.BASIS == "Forecast")
                        & assistance.SPEND_SOURCE.str.startswith("Treasury")]
        cols = st.columns(4)
        cols[0].metric("Total assistance, %s" % y, "${:,.0f}m".format(tot),
                       help="Latest year covered by both Treasury and MSD.")
        biggest = year_rows.nlargest(1, "AMOUNT_NZD_MILLION")
        if not biggest.empty:
            b = biggest.iloc[0]
            cols[1].metric("Largest programme", b.PROGRAMME,
                           delta="${:,.0f}m".format(b.AMOUNT_NZD_MILLION), delta_color="off")
        if not fc.empty:
            fy = fc.PERIOD.max()
            base = tsy[tsy.PERIOD == y].AMOUNT_NZD_MILLION.sum()
            ftot = fc[fc.PERIOD == fy].AMOUNT_NZD_MILLION.sum()
            cols[2].metric("Benefit forecast, %s" % fy, "${:,.0f}m".format(ftot),
                           delta="{:+.1f}% on {}".format(100 * (ftot - base) / base, y),
                           help="Treasury benefit lines only; student support is not forecast.")
        cols[3].metric("Programmes", "{:,}".format(year_rows.PROGRAMME.nunique()))


def render_assistance_people():
    """People supported by programme, one line each, deliberately never stacked.

    The caption is part of the visual rather than decoration: these populations
    overlap — one person can hold a student loan and a main benefit — so an area
    chart or a total would be wrong.
    """
    st.markdown("##### People supported")
    people = get_people_supported(df_db_schema)
    if people.empty:
        st.info("No recipient series available.")
        return
    fig = px.line(people.sort_values("YEAR"), x="YEAR", y="RECIPIENTS", color="PROGRAMME",
                  color_discrete_sequence=px.colors.qualitative.Safe, markers=True,
                  labels={"YEAR": "", "RECIPIENTS": "People", "PROGRAMME": ""})
    fig.update_layout(height=420, hovermode="x unified", legend=dict(orientation="h", y=-0.25),
                      margin=dict(t=10, b=10, l=0, r=0))
    fig.update_yaxes(tickformat=",")
    st.plotly_chart(fig, use_container_width=True)
    st.caption(
        "These series are shown together but must not be added. Benefits and "
        "pensions are counts at the December quarter; StudyLink counts everyone "
        "supported at any point in the year; and one person can hold a student "
        "loan and a main benefit at the same time."
    )


def render_assistance_detail(assistance):
    """Full programme-level spend table, with an Excel export and a sourcing caption."""
    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📋 Assistance spending by programme")
    with dl:
        st.download_button(
            "📥 Excel",
            data=build_styled_excel(assistance, "Government assistance by programme",
                                    "Assistance"),
            file_name="msd_all_assistance.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_assist_xlsx", type="primary")
    st.dataframe(assistance, use_container_width=True, hide_index=True, height=360)
    st.caption(
        "Spending is in NZ$ million. Treasury lines come from the Budget Economic "
        "and Fiscal Update core Crown expense tables and reconcile exactly to the "
        "published benefit expenses total; student support comes from MSD."
    )


# ====================STATIC_METHODS====================
# Shared helpers that draw nothing on their own.
def _safe_filename(text):
    """Trim any label into something safe to use as a download filename."""
    cleaned = "".join(c for c in str(text) if c.isalnum() or c in " _-").strip().replace(" ", "_")
    return cleaned[:40] or "results"


def build_styled_excel(df, title, sheet_name="Results"):
    """Generic styled in-memory Excel export for any dataframe.

    Every table on every tab downloads through this one function, so exports
    look the same wherever they came from: navy title bar, blue header row,
    banded rows, dates and numbers typed rather than stringified, frozen panes
    and an autofilter on the header.
    """
    cols = list(df.columns)
    ncols = max(len(cols), 1)
    wrap_cols = {"SOURCE_URL", "SECTION", "REASON", "MEASURE_GROUP"}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.merge_cells("A1:%s1" % get_column_letter(ncols))
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    hdr_fill = PatternFill("solid", start_color="2E86AB")
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_border = Border(bottom=Side(style="medium", color="1A3A5C"))
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=str(col))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = hdr_border
    ws.row_dimensions[2].height = 18

    is_dt = {c: pd.api.types.is_datetime64_any_dtype(df[c]) for c in cols}
    is_num = {c: pd.api.types.is_numeric_dtype(df[c]) for c in cols}
    alt_fill = PatternFill("solid", start_color="EBF5FB")
    thin_border = Border(bottom=Side(style="thin", color="D5D8DC"))

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        row_fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(cols, start=1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            if (pd.isna(val) if not isinstance(val, str) else val == ""):
                cell.value = None
            elif is_dt[col]:
                try:
                    cell.value = pd.to_datetime(val).to_pydatetime()
                    cell.number_format = "DD-MMM-YY"
                except Exception:
                    cell.value = str(val)
            elif is_num[col]:
                try:
                    fv = float(val)
                    cell.value = int(fv) if fv.is_integer() else round(fv, 4)
                except Exception:
                    cell.value = str(val)
            else:
                cell.value = str(val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    for ci, col in enumerate(cols, start=1):
        try:
            sample = df[col].head(200).astype(str).map(len).max()
            maxlen = max(len(str(col)), int(sample) if pd.notna(sample) else 0)
        except Exception:
            maxlen = len(str(col))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 2, 10), 45)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:%s2" % get_column_letter(ncols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ====================MAIN====================
def main():
    """Page assembly: banner, then sidebar controls, then tabs, then attribution.

    Deliberately thin. Every decision about what is drawn lives in the render_*
    functions above, so this reads as the page's table of contents.
    """
    if not os.path.exists(DB_PATH):
        st.error("Data file not found at %s." % DB_PATH)
        return
    render_header()
    filters = render_sidebar()
    render_main_tabs(filters)
    render_attribution()


if __name__ == "__main__":
    main()
